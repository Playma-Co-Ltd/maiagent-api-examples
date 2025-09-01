import os
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import MaiAgentHelper

# 設定
API_KEY = '0yZD0JZY.PhUtANQUat5X5sWEav2GQgVkXJ2uR34D'
KNOWLEDGE_BASE_ID = '1491857c-aefc-464b-a4c6-b6494a8172d9'
EXCEL_FILE_PATH = './downloads/knowledge_base_faqs_6b11f385-fc69-480e-982f-85a395954d5c_20250824_223813.xlsx'  # 例如: 'faqs_template.xlsx'

assert API_KEY != '<your-api-key>', 'Please set your API key'
assert KNOWLEDGE_BASE_ID != '<your-knowledge-base-id>', 'Please set your knowledge base id'
assert EXCEL_FILE_PATH != '<your-excel-file-path>', 'Please set your Excel file path'


def main():
    """
    從 Excel 檔案批量上傳 FAQ
    
    支援從標準格式的 Excel 檔案讀取 FAQ 數據並批量上傳到 MaiAgent
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安裝 openpyxl 套件才能讀取 Excel 檔案")
        print("請執行：pip install openpyxl")
        return

    maiagent_helper = MaiAgentHelper(API_KEY)

    print("=" * 60)
    print("MaiAgent FAQ Excel 批量上傳工具")
    print("=" * 60)
    
    try:
        # 1. 檢查檔案是否存在
        if not os.path.exists(EXCEL_FILE_PATH):
            print(f"❌ Excel 檔案不存在：{EXCEL_FILE_PATH}")
            print("請確認檔案路徑正確，或使用 faq_excel_template.xlsx 模板")
            return
        
        # 2. 讀取 Excel 檔案
        print(f"\n📖 正在讀取 Excel 檔案：{EXCEL_FILE_PATH}")
        faq_data = read_excel_file(EXCEL_FILE_PATH)
        
        if not faq_data:
            print("❌ Excel 檔案中沒有找到有效的 FAQ 數據")
            return
        
        print(f"✅ 成功讀取 {len(faq_data)} 個 FAQ")
        
        # 3. 顯示預覽
        display_faq_preview(faq_data)
        
        # 4. 確認上傳
        if not confirm_upload(len(faq_data)):
            print("操作已取消")
            return
        
        # 5. 批量上傳 FAQ
        results = upload_faqs_batch(maiagent_helper, faq_data, KNOWLEDGE_BASE_ID)
        
        # 6. 顯示結果
        display_upload_results(results)
        
        # 7. 生成報告
        generate_upload_report(results, EXCEL_FILE_PATH)
        
    except Exception as e:
        print(f"❌ 批量上傳失敗：{e}")


def read_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """讀取 Excel 檔案並解析 FAQ 數據"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安裝 openpyxl 套件")
    
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    
    # 嘗試找到 FAQ 工作表
    worksheet = None
    possible_names = ['FAQ List', 'FAQs', 'FAQ', 'Sheet1', workbook.sheetnames[0]]
    
    for name in possible_names:
        if name in workbook.sheetnames:
            worksheet = workbook[name]
            break
    
    if worksheet is None:
        raise ValueError("找不到包含 FAQ 數據的工作表")
    
    # 讀取標題列
    headers = []
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for cell in first_row:
        if cell:
            headers.append(str(cell).strip().lower())
    
    print(f"📋 找到的欄位：{headers}")
    
    # 檢查必要欄位
    required_fields = ['question', 'answer']
    question_col = find_column_index(headers, ['question', 'q', '問題', '問'])
    answer_col = find_column_index(headers, ['answer', 'a', '答案', '答', '回答'])
    
    if question_col is None or answer_col is None:
        raise ValueError(f"找不到必要的欄位。需要包含 'Question' 和 'Answer' 欄位。\n找到的欄位：{headers}")
    
    # 找出可選欄位
    labels_col = find_column_index(headers, ['labels', 'tags', 'label', 'tag', '標籤', '標記'])
    
    print(f"✅ 欄位對應：Question={question_col+1}, Answer={answer_col+1}, Labels={labels_col+1 if labels_col else 'N/A'}")
    
    # 讀取數據
    faq_data = []
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= max(question_col, answer_col):
            continue
            
        question = str(row[question_col]).strip() if row[question_col] else ''
        answer = str(row[answer_col]).strip() if row[answer_col] else ''
        
        # 跳過空白行
        if not question and not answer:
            continue
        
        if not question or not answer:
            print(f"⚠️  第 {row_num} 行資料不完整，已跳過")
            continue
        
        # 處理標籤
        labels = []
        if labels_col is not None and labels_col < len(row) and row[labels_col]:
            labels_str = str(row[labels_col]).strip()
            if labels_str:
                # 支援多種分隔符
                for sep in [',', ';', '|', '，', '；']:
                    if sep in labels_str:
                        labels = [label.strip() for label in labels_str.split(sep) if label.strip()]
                        break
                if not labels:
                    labels = [labels_str]
        
        faq_data.append({
            'question': question,
            'answer': answer,
            'labels': labels,
            'row_number': row_num
        })
    
    workbook.close()
    return faq_data


def find_column_index(headers: List[str], possible_names: List[str]) -> Optional[int]:
    """找出欄位在標題列中的索引"""
    for i, header in enumerate(headers):
        for name in possible_names:
            if name.lower() in header.lower():
                return i
    return None


def display_faq_preview(faq_data: List[Dict[str, Any]]):
    """顯示 FAQ 預覽"""
    print("\n📝 FAQ 預覽（前 3 個）：")
    for i, faq in enumerate(faq_data[:3], 1):
        question = faq['question']
        answer = faq['answer']
        labels = faq.get('labels', [])
        
        print(f"   {i}. Q: {question[:80]}{'...' if len(question) > 80 else ''}")
        print(f"      A: {answer[:80]}{'...' if len(answer) > 80 else ''}")
        if labels:
            print(f"      標籤: {', '.join(labels)}")
        print(f"      來源: 第 {faq['row_number']} 行")
        print()


def confirm_upload(total_count: int) -> bool:
    """確認是否上傳"""
    print(f"\n❓ 確認要上傳 {total_count} 個 FAQ 嗎？")
    print("注意：這個操作會調用 API 新增 FAQ，可能會消耗 API 配額")
    
    try:
        choice = input("繼續上傳？(y/N): ").strip().lower()
        return choice in ['y', 'yes', 'Y', 'YES']
    except KeyboardInterrupt:
        return False


def upload_faqs_batch(maiagent_helper: MaiAgentHelper, faq_data: List[Dict[str, Any]], knowledge_base_id: str) -> Dict[str, Any]:
    """批量上傳 FAQ"""
    results = {
        'success': [],
        'failed': [],
        'total': len(faq_data),
        'start_time': datetime.now()
    }
    
    print(f"\n🚀 開始批量上傳 {len(faq_data)} 個 FAQ...")
    print("=" * 50)
    
    for i, faq in enumerate(faq_data, 1):
        try:
            print(f"[{i:3d}/{len(faq_data)}] 正在上傳: {faq['question'][:50]}{'...' if len(faq['question']) > 50 else ''}")
            
            # 調用 create_knowledge_base_faq API
            response = create_knowledge_base_faq(
                maiagent_helper,
                knowledge_base_id=knowledge_base_id,
                question=faq['question'],
                answer=faq['answer'],
                labels=faq.get('labels', [])
            )
            
            if response and 'id' in response:  # 成功創建
                results['success'].append({
                    'row_number': faq['row_number'],
                    'question': faq['question'],
                    'answer': faq['answer'],
                    'labels': faq.get('labels', []),
                    'faq_id': response['id']
                })
                print(f"         ✅ 成功 (ID: {response['id'][:8]}...)")
            else:
                results['failed'].append({
                    'row_number': faq['row_number'],
                    'question': faq['question'],
                    'error': 'API 調用失敗',
                    'details': 'API 返回空響應或缺少 ID'
                })
                print(f"         ❌ 失敗")
                
        except Exception as e:
            results['failed'].append({
                'row_number': faq['row_number'],
                'question': faq['question'],
                'error': str(e),
                'details': 'API 調用異常'
            })
            print(f"         ❌ 失敗: {e}")
    
    results['end_time'] = datetime.now()
    results['duration'] = results['end_time'] - results['start_time']
    
    return results


def create_knowledge_base_faq(maiagent_helper: MaiAgentHelper, knowledge_base_id: str, question: str, answer: str, labels: List[str] = None) -> Dict[str, Any]:
    """新增單個知識庫 FAQ"""
    try:
        # 處理標籤格式
        formatted_labels = []
        if labels:
            # 由於我們無法直接獲取標籤ID，先只使用標籤名稱
            # 實際使用時可能需要先查詢或創建標籤
            for label_name in labels:
                if label_name.strip():
                    formatted_labels.append({"name": label_name.strip()})
        
        response = maiagent_helper.create_knowledge_base_faq(
            knowledge_base_id=knowledge_base_id,
            question=question,
            answer=answer,
            labels=formatted_labels if formatted_labels else None
        )
        return response
    except Exception as e:
        raise Exception(f"知識庫 FAQ 創建失敗: {e}")


def display_upload_results(results: Dict[str, Any]):
    """顯示上傳結果"""
    success_count = len(results['success'])
    failed_count = len(results['failed'])
    total_count = results['total']
    duration = results['duration']
    
    print("\n" + "=" * 50)
    print("📊 上傳結果統計")
    print("=" * 50)
    print(f"總計 FAQ：{total_count}")
    print(f"成功上傳：{success_count} ✅")
    print(f"上傳失敗：{failed_count} ❌")
    print(f"成功率：{(success_count/total_count*100):.1f}%")
    print(f"耗時：{duration.total_seconds():.1f} 秒")
    
    if failed_count > 0:
        print(f"\n❌ 失敗的 FAQ：")
        for failed in results['failed']:
            print(f"   第 {failed['row_number']} 行: {failed['question'][:60]}...")
            print(f"      錯誤: {failed['error']}")


def generate_upload_report(results: Dict[str, Any], source_file: str):
    """生成上傳報告"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"faq_upload_report_{timestamp}.txt"
    
    # 確保 downloads 目錄存在
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    downloads_dir = os.path.join(script_dir, "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    
    report_path = os.path.join(downloads_dir, report_filename)
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("MaiAgent FAQ 批量上傳報告\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"來源檔案：{source_file}\n")
        f.write(f"上傳時間：{results['start_time'].strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"完成時間：{results['end_time'].strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"總耗時：{results['duration'].total_seconds():.1f} 秒\n\n")
        
        f.write("統計資訊：\n")
        f.write(f"  總計 FAQ：{results['total']}\n")
        f.write(f"  成功上傳：{len(results['success'])}\n")
        f.write(f"  上傳失敗：{len(results['failed'])}\n")
        f.write(f"  成功率：{(len(results['success'])/results['total']*100):.1f}%\n\n")
        
        if results['success']:
            f.write("成功上傳的 FAQ：\n")
            f.write("-" * 40 + "\n")
            for i, item in enumerate(results['success'], 1):
                f.write(f"{i}. 第 {item['row_number']} 行\n")
                f.write(f"   問題：{item['question']}\n")
                f.write(f"   答案：{item['answer'][:100]}{'...' if len(item['answer']) > 100 else ''}\n")
                if item['labels']:
                    f.write(f"   標籤：{', '.join(item['labels'])}\n")
                f.write("\n")
        
        if results['failed']:
            f.write("失敗的 FAQ：\n")
            f.write("-" * 40 + "\n")
            for i, item in enumerate(results['failed'], 1):
                f.write(f"{i}. 第 {item['row_number']} 行\n")
                f.write(f"   問題：{item['question']}\n")
                f.write(f"   錯誤：{item['error']}\n")
                f.write(f"   詳情：{item['details']}\n")
                f.write("\n")
    
    print(f"\n📄 詳細報告已保存到：{report_path}")


if __name__ == '__main__':
    main()
