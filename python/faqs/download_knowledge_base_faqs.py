import csv
import json
import os
import sys
from datetime import datetime
from typing import Any, Dict, List

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import MaiAgentHelper

# 設定
API_KEY = '0yZD0JZY.PhUtANQUat5X5sWEav2GQgVkXJ2uR34D'
KNOWLEDGE_BASE_ID = '6b11f385-fc69-480e-982f-85a395954d5c'   # 你的知識庫 ID

assert API_KEY != '<your-api-key>', 'Please set your API key'
assert KNOWLEDGE_BASE_ID != '<your-knowledge-base-id>', 'Please set your knowledge base id'


def main():
    """
    下載知識庫 FAQ 範例
    
    展示如何使用 MaiAgent API 下載知識庫中的所有 FAQ
    並支援多種匯出格式：JSON、CSV、Excel
    """
    maiagent_helper = MaiAgentHelper(API_KEY)

    try:
        print("=" * 60)
        print("MaiAgent 知識庫 FAQ 下載工具")
        print("=" * 60)
        
        # 1. 獲取所有 FAQ（支持分頁）
        print(f"\n📋 正在從知識庫 {KNOWLEDGE_BASE_ID} 獲取 FAQ 列表...")
        print("⚠️  如果 FAQ 數量很多，這可能需要一些時間...")
        faqs_response = maiagent_helper.list_all_knowledge_base_faqs(KNOWLEDGE_BASE_ID)
        
        if 'results' not in faqs_response:
            print("❌ 無法獲取 FAQ 列表或知識庫中沒有 FAQ")
            return
        
        faq_list = faqs_response['results']
        total_count = faqs_response.get('count', len(faq_list))
        
        print(f"\n✅ 成功獲取所有 FAQ！")
        print(f"   總計：{total_count} 個 FAQ")
        
        if not faq_list:
            print("知識庫中沒有 FAQ 可下載")
            return

        # 2. 顯示 FAQ 統計資訊
        display_faq_statistics(faq_list)
        
        # 3. 選擇導出格式
        export_format = choose_export_format()
        if not export_format:
            print("操作已取消")
            return
            
        # 4. 導出 FAQ
        export_faqs(faq_list, KNOWLEDGE_BASE_ID, export_format)
            
    except Exception as e:
        print(f"❌ FAQ 下載失敗：{e}")


def display_faq_statistics(faq_list: List[Dict[str, Any]]):
    """顯示 FAQ 統計資訊"""
    print("\n📊 FAQ 統計資訊：")
    print(f"   總 FAQ 數量: {len(faq_list)}")
    
    if len(faq_list) > 100:
        print(f"   📄 資料量較大，已自動處理分頁獲取")
    
    # 統計有標籤的 FAQ
    faqs_with_labels = [faq for faq in faq_list if faq.get('labels')]
    print(f"   有標籤的 FAQ: {len(faqs_with_labels)}")
    
    # 統計標籤
    all_labels = []
    for faq in faq_list:
        labels = faq.get('labels', [])
        for label in labels:
            if isinstance(label, dict):
                all_labels.append(label.get('name', ''))
            else:
                all_labels.append(str(label))
    
    unique_labels = set(filter(None, all_labels))
    if unique_labels:
        print(f"   標籤種類: {len(unique_labels)} ({', '.join(list(unique_labels)[:5])}{'...' if len(unique_labels) > 5 else ''})")
    
    # 預覽前 3 個 FAQ
    print("\n📝 FAQ 預覽：")
    for i, faq in enumerate(faq_list[:3], 1):
        question = faq.get('question', '').strip()
        answer = faq.get('answer', '').strip()
        print(f"   {i}. Q: {question[:50]}{'...' if len(question) > 50 else ''}")
        print(f"      A: {answer[:50]}{'...' if len(answer) > 50 else ''}")
        print(f"      ID: {faq.get('id')}")
        if faq.get('labels'):
            labels_str = ', '.join([
                label.get('name', str(label)) if isinstance(label, dict) else str(label)
                for label in faq.get('labels', [])
            ])
            print(f"      標籤: {labels_str}")
        print()


def choose_export_format() -> str:
    """選擇導出格式"""
    print("\n📤 選擇導出格式：")
    print("1. JSON 格式 (.json) - 完整的結構化數據")
    print("2. CSV 格式 (.csv) - 表格數據，適合 Excel 開啟")
    print("3. Excel 格式 (.xlsx) - Excel 工作簿")
    print("4. 全部格式")
    print("5. 取消")
    
    try:
        choice = input("\n請輸入選擇 (1-5): ").strip()
    except KeyboardInterrupt:
        print("\n操作已取消")
        return None
    
    format_map = {
        '1': 'json',
        '2': 'csv', 
        '3': 'excel',
        '4': 'all',
        '5': None
    }
    
    return format_map.get(choice)


def export_faqs(faq_list: List[Dict[str, Any]], knowledge_base_id: str, export_format: str):
    """導出 FAQ 到指定格式"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = f"knowledge_base_faqs_{knowledge_base_id}_{timestamp}"
    
    # 確保下載目錄存在（使用絕對路徑）
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 回到 python/ 目錄
    downloads_dir = os.path.join(script_dir, "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    
    print(f"📂 下載目錄：{downloads_dir}")
    
    exported_files = []
    
    try:
        if export_format in ['json', 'all']:
            json_file = export_to_json(faq_list, downloads_dir, base_filename, knowledge_base_id)
            if json_file:
                exported_files.append(json_file)
                
        if export_format in ['csv', 'all']:
            csv_file = export_to_csv(faq_list, downloads_dir, base_filename)
            if csv_file:
                exported_files.append(csv_file)
                
        if export_format in ['excel', 'all']:
            excel_file = export_to_excel(faq_list, downloads_dir, base_filename, knowledge_base_id)
            if excel_file:
                exported_files.append(excel_file)
        
        # 顯示結果
        if exported_files:
            print(f"\n✅ FAQ 導出成功！")
            print(f"📁 導出位置：{downloads_dir}")
            for file_path in exported_files:
                file_size = os.path.getsize(file_path)
                print(f"   📄 {os.path.basename(file_path)} ({format_file_size(file_size)})")
                
            print(f"\n💡 提示：您可以直接訪問上述目錄查看下載的檔案")
            
            print(f"\n📊 導出統計：")
            print(f"   總 FAQ 數量：{len(faq_list)}")
            print(f"   導出檔案數：{len(exported_files)}")
        else:
            print("❌ 沒有檔案被導出")
            
    except Exception as e:
        print(f"❌ 導出過程中發生錯誤：{e}")


def export_to_json(faq_list: List[Dict[str, Any]], downloads_dir: str, base_filename: str, knowledge_base_id: str) -> str:
    """導出為 JSON 格式"""
    print("\n📝 正在導出 JSON 格式...")
    
    json_filename = f"{base_filename}.json"
    json_path = os.path.join(downloads_dir, json_filename)
    
    # 準備導出數據
    export_data = {
        "knowledge_base_id": knowledge_base_id,
        "export_time": datetime.now().isoformat(),
        "total_faqs": len(faq_list),
        "faqs": faq_list
    }
    
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON 導出完成：{json_filename}")
    return json_path


def export_to_csv(faq_list: List[Dict[str, Any]], downloads_dir: str, base_filename: str) -> str:
    """導出為 CSV 格式"""
    print("\n📊 正在導出 CSV 格式...")
    
    csv_filename = f"{base_filename}.csv"
    csv_path = os.path.join(downloads_dir, csv_filename)
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # 寫入標題列
        writer.writerow(['ID', 'Question', 'Answer', 'Labels', 'Created_At', 'Updated_At'])
        
        # 寫入 FAQ 數據
        for faq in faq_list:
            labels_str = ''
            if faq.get('labels'):
                labels = [
                    label.get('name', str(label)) if isinstance(label, dict) else str(label)
                    for label in faq.get('labels', [])
                ]
                labels_str = ', '.join(labels)
            
            writer.writerow([
                faq.get('id', ''),
                faq.get('question', ''),
                faq.get('answer', ''),
                labels_str,
                faq.get('created_at', ''),
                faq.get('updated_at', '')
            ])
    
    print(f"✅ CSV 導出完成：{csv_filename}")
    return csv_path


def export_to_excel(faq_list: List[Dict[str, Any]], downloads_dir: str, base_filename: str, knowledge_base_id: str) -> str:
    """導出為 Excel 格式"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 需要安裝 openpyxl 套件才能導出 Excel 格式")
        print("請執行：pip install openpyxl")
        return None
    
    print("\n📈 正在導出 Excel 格式...")
    
    excel_filename = f"{base_filename}.xlsx"
    excel_path = os.path.join(downloads_dir, excel_filename)
    
    # 創建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FAQ List"
    
    # 設定樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 寫入標題列
    headers = ['ID', 'Question', 'Answer', 'Labels', 'Created At', 'Updated At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 寫入 FAQ 數據
    for row, faq in enumerate(faq_list, 2):
        ws.cell(row=row, column=1, value=faq.get('id', ''))
        ws.cell(row=row, column=2, value=faq.get('question', ''))
        ws.cell(row=row, column=3, value=faq.get('answer', ''))
        
        # 處理標籤
        labels_str = ''
        if faq.get('labels'):
            labels = [
                label.get('name', str(label)) if isinstance(label, dict) else str(label)
                for label in faq.get('labels', [])
            ]
            labels_str = ', '.join(labels)
        ws.cell(row=row, column=4, value=labels_str)
        
        ws.cell(row=row, column=5, value=faq.get('created_at', ''))
        ws.cell(row=row, column=6, value=faq.get('updated_at', ''))
    
    # 調整欄寬
    column_widths = [36, 50, 80, 30, 20, 20]  # ID, Question, Answer, Labels, Created, Updated
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 創建統計工作表
    stats_ws = wb.create_sheet("Statistics")
    stats_data = [
        ["Knowledge Base ID", knowledge_base_id],
        ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total FAQs", len(faq_list)],
        ["FAQs with Labels", len([faq for faq in faq_list if faq.get('labels')])],
    ]
    
    for row, (key, value) in enumerate(stats_data, 1):
        stats_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        stats_ws.cell(row=row, column=2, value=str(value))
    
    stats_ws.column_dimensions['A'].width = 20
    stats_ws.column_dimensions['B'].width = 40
    
    # 保存工作簿
    wb.save(excel_path)
    
    print(f"✅ Excel 導出完成：{excel_filename}")
    return excel_path


def format_file_size(size_bytes):
    """格式化檔案大小"""
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB"]
    size_bytes = int(size_bytes)
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"


if __name__ == '__main__':
    main()
